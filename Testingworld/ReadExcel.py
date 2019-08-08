import openpyxl

#Load workbook

wk=openpyxl.load_workbook('C:\\Users\\Bilal\\PycharmProjects\\TestingAutomation\\ExcelRead\\Testcases.xlsx', 'read_only')

print(wk.sheetnames)

print("Active sheet is"+wk.active.title)

#Create object of any sheet on which you want to work

sh = wk['Sheet1']
print(sh.title)

#Fetch the value
print(sh['A2'].value)
print(sh['B3'].value)
print(sh['C4'].value)

# #Creating cell object
#
# cl =sh.cell(4,3)
# print(cl.value)
#
# #Find Rows Having Data
# rows= sh.max_row
# columns =sh.max_column
#
# print("Total Rows are - " + str(rows))
# print("Total Columns are - " + str(columns))
#
# for i in range(1,rows+1):
#     for j in range(1, columns + 1):
#         c = sh.cell(i,j)
#         print(c.value)
#
#         for r in sh['A1':'C4']:
#             for c in r:
#                 print(c.value)